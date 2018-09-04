from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet('My_sheet1')
ws2 = wb.create_sheet('MySheet2')
ws.title = 'New Title---'
ws2.sheet_properties.tabColor = '1072BA'
print(wb.sheetnames)
for sheet in wb:
    print('for + '+sheet.title)
source = wb.active
target = wb.copy_worksheet(source)
A4 = ws['A4']
A4 = 44444
cell = ws.cell(row=4, column=2, value=10)

'''
Because of this feature, scrolling through cells instead of accessing them directly will create them all in memory, even if you don’t assign them a value.
'''

for i in range(1, 101):
    for j in range(1, 101):
        a=ws.cell(row=i, column=j)

cell_range=ws['A1':'C2']
colC=ws['C']
col_range=ws['C:D']
row10=ws[10]
row_range=ws[5:10]

for row in ws.iter_rows(min_row=1,max_col=3, max_row=2):
    for cell in row:
        print(cell)

ws['F7']='hello world'
print(ws['F7'].value)

