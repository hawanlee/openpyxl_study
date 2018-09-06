import datetime
from openpyxl import Workbook

wb=Workbook()
ws=wb.active
# set date using a Python datetime
ws['A1']=datetime.datetime(2010,7,10)
print(ws['A1'].number_format)

# enable type inerence on a case-by-case basis
wb.guess_types=True
ws['B1']='3.14%'
print(ws['B1'].value)
print(ws['B1'].number_format)
wb.guess_types=False
print(ws['B1'].value)
print(ws['B1'].number_format)

# using formula
ws['C1']='=SUM(1,1)'

# merge/unmerge cells
ws.merge_cells('A3:D3')
ws.merge_cells('A4:D4')
ws.unmerge_cells('A4:D4')
# or
ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=7)

# insert an image
from openpyxl.drawing.image import Image

ws['A7']='You should see threelogos below'
img=Image('cell.png')
ws.add_image(img,'A1')

# fold

import openpyxl
ws5=wb.create_sheet()
ws.column_dimensions.group('G','H',hidden=True)

wb.save('3_formula.xlsx')