#! /usr/bin/python
# coding=utf-8

import csv
import datetime
from openpyxl import load_workbook
import time
from copy import copy
from openpyxl.styles import Border, Side, Font, Alignment

with open('Task.csv',encoding='utf-8') as csvfile:
    reader=csv.reader(csvfile)
    rows=[row for row in reader]
print(len(rows))

reviewTime = []
for i in range(0, len(rows)):
    reviewTime.append((rows[i][8])[:-3])
    # 切除后三位字符
    if i != 0:
        reviewTime[i] = datetime.datetime.strptime(
            reviewTime[i], '%Y/%m/%d %H:%M:%S')
        # 将字符串转换为标准日期格式
print(reviewTime)

topic = []
for i in range(0, len(rows)):
    topic.append(rows[i][3])
print(topic)

pipelineStep = []
for i in range(0, len(rows)):
    pipelineStep.append(rows[i][4])
print(pipelineStep)

reviewPurpose = []
for i in range(0, len(rows)):
    reviewPurpose.append(rows[i][10])
print(reviewPurpose)

presenter = []
for i in range(0, len(rows)):
    presenter.append(rows[i][5])
print(presenter)

reviewSite = []
for i in range(0, len(rows)):
    reviewSite.append(rows[i][9])
print(reviewSite)

'''
a=datetime.datetime.strptime('2018/09/06 09:00:00', '%Y/%m/%d %H:%M:%S')
b=datetime.datetime.strptime('2018/09/06 09:05:00', '%Y/%m/%d %H:%M:%S')
print(int(((b-a).seconds)/60))
# 输出时间差
'''

minutes = []
for i in range(0, len(rows)-2):
    a1 = reviewTime[i+1]
    a2 = reviewTime[i+2]
    minutes.append(str(int(((a2-a1).seconds)/60))+' min')
minutes.append('TBD')
print(minutes)

te=load_workbook('TemplateEDR.xlsx')
ws=te.active
print(ws['A5'].value)

for i in range(0, len(rows)-1):
    column_A = 'A'+str(i+5)
    ttt = datetime.datetime.strftime(reviewTime[i+1], '%H:%M')
    # 将标准日期格式转换为字符串
    ws[column_A].value = ttt
    column_B = 'B'+str(i+5)
    ws[column_B].value = minutes[i]
    column_C = 'C'+str(i+5)
    ws[column_C] = topic[i+1]
    column_D = 'D'+str(i+5)
    ws[column_D] = pipelineStep[i+1]
    ws['E'+str(i+5)] = reviewPurpose[i+1]
    ws['F'+str(i+5)] = presenter[i+1]
    ws['G'+str(i+5)] = reviewSite[i+1]

for i in range(0,len(rows)-1):
    for j in range(0,7):
        column=str(chr(65+j)+str(i+5))
        border1 = Border(left=Side(style='thin', color='FF000000'), right=Side  (style='thin', color='FF000000'),
                        top=Side(style='thin', color='FF000000'), bottom=Side  (style='thin', color='FF000000'))
        ws[column].border = border1
        font1=Font(name='Century Gothic',size=11)
        ws[column].font=font1
        alignment1=Alignment(wrap_text=True)
        ws[column].alignment=alignment1

ws.page_setup.fitToHeight = 1
ws.page_setup.fitToWidth = 1

# 设置表头周数
dateTitle=int(datetime.datetime.strftime(reviewTime[2],'%U'))+1
print(dateTitle)
ws['A1'].value = ws['A1'].value[:-2]+str(dateTitle)
print(ws['A1'].value)

# 设置表头会议时间
cfTime=datetime.datetime.strftime(reviewTime[1],'%d%b, %Y, %H:%M')
ws['A2'].value=cfTime

date1=datetime.datetime.strftime(reviewTime[1], '%Y%m%d')

te.save('EDR review_'+str(date1)+'.xlsx')

